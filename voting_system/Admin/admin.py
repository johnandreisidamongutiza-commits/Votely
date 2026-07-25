from flask import Blueprint, render_template, session, redirect, url_for, request, jsonify, send_from_directory, send_file
import mysql.connector
import os
import random
import io
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from voting_system import db_config
import openpyxl

UPLOAD_FOLDER = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

admin = Blueprint('admin', __name__, template_folder='templates', static_folder='static', static_url_path='/admin/static')

SESSION_TIMEOUT = timedelta(minutes=30)


def _log(cursor, user_id, action, election_id, details):
    """Insert an activity log only if the user still exists in the DB."""
    cursor.execute("SELECT id FROM users WHERE id=%s", (user_id,))
    if cursor.fetchone():
        cursor.execute(
            "INSERT INTO activity_logs (user_id, action, election_id, details) VALUES (%s, %s, %s, %s)",
            (user_id, action, election_id, details)
        )

@admin.before_request
def check_admin_session():
    if 'admin_id' not in session:
        return
    last = session.get('admin_last_activity')
    if last and datetime.now() - datetime.fromisoformat(last) > SESSION_TIMEOUT:
        session.pop('admin_id', None)
        session.pop('admin_name', None)
        session.pop('admin_last_activity', None)
        return redirect(url_for('auth.login') + '?error=You have been logged out due to inactivity.')
    session['admin_last_activity'] = datetime.now().isoformat()

def update_election_statuses():
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute("UPDATE elections SET status='active' WHERE status='draft' AND start_date <= %s AND end_date >= %s", (now, now))
    cursor.execute("UPDATE elections SET status='ended' WHERE status='active' AND end_date < %s", (now,))
    connect.commit()
    cursor.close()
    connect.close()

@admin.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@admin.route('/dashboard', defaults={'section': 'elections'})
@admin.route('/dashboard/<section>')
def dashboard(section):
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    valid_sections = ['elections', 'register-voter', 'candidates', 'results', 'logs']
    if section not in valid_sections:
        return redirect(url_for('admin.elections'))
    update_election_statuses()
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT * FROM elections WHERE status IN ('draft','active') ORDER BY created_at DESC")
    ongoing_elections = cursor.fetchall()
    cursor.execute("SELECT * FROM elections WHERE status='ended' AND is_archived=0 ORDER BY end_date DESC")
    history_elections = cursor.fetchall()
    cursor.execute("SELECT * FROM elections WHERE is_archived=1 ORDER BY end_date DESC")
    archived_elections = cursor.fetchall()
    inactive_voters = []
    elections = []
    if section == 'register-voter':
        cursor.execute("SELECT id, user_id, firstname, lastname, email FROM users WHERE status='inactive' AND role='user' ORDER BY firstname")
        inactive_voters = cursor.fetchall()
    if section in ('logs', 'candidates', 'results'):
        cursor.execute("SELECT * FROM elections ORDER BY created_at DESC")
        elections = cursor.fetchall()
    cursor.close()
    connect.close()
    templates = {
        'elections': 'admin_elections.html',
        'register-voter': 'admin_register_voter.html',
        'candidates': 'admin_candidates.html',
        'results': 'admin_results.html',
        'logs': 'admin_logs.html'
    }
    return render_template(templates[section], username=session.get('admin_name'), ongoing_elections=ongoing_elections, history_elections=history_elections, archived_elections=archived_elections, inactive_voters=inactive_voters, elections=elections, active_page=section.replace('-', '_'))

@admin.route('/elections')
def elections():
    return redirect(url_for('admin.dashboard', section='elections'))

@admin.route('/register-voter-page')
def register_voter_page():
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT id, user_id, firstname, lastname, email FROM users WHERE status='inactive' AND role='user' ORDER BY firstname")
    inactive_voters = cursor.fetchall()
    cursor.close()
    connect.close()
    return render_template('admin_register_voter.html', username=session.get('admin_name'), inactive_voters=inactive_voters, active_page='register_voter')

@admin.route('/candidates-page')
def candidates():
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    update_election_statuses()
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT * FROM elections WHERE status='draft' ORDER BY created_at DESC")
    elections = cursor.fetchall()
    cursor.close()
    connect.close()
    return render_template('admin_candidates.html', username=session.get('admin_name'), elections=elections, active_page='candidates')

@admin.route('/results-page')
def results_page():
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    update_election_statuses()
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT * FROM elections ORDER BY created_at DESC")
    elections = cursor.fetchall()
    cursor.close()
    connect.close()
    return render_template('admin_results.html', username=session.get('admin_name'), elections=elections, active_page='results')

@admin.route('/logs-page')
def logs_page():
    return redirect(url_for('admin.dashboard', section='logs'))

@admin.route('/create-election', methods=['POST'])
def create_election():
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    if request.method == 'POST':
        title       = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        start_date  = request.form.get('start_date', '')
        end_date    = request.form.get('end_date', '')

        if not title:
            return jsonify({'success': False, 'error': 'Title is required.'})
        if len(title) < 5:
            return jsonify({'success': False, 'error': 'Title must be at least 5 characters.'})
        if len(title) > 150:
            return jsonify({'success': False, 'error': 'Title cannot exceed 150 characters.'})
        if len(description) > 300:
            return jsonify({'success': False, 'error': 'Description cannot exceed 300 characters.'})
        if not start_date or not end_date:
            return jsonify({'success': False, 'error': 'Start and end dates are required.'})

        from datetime import datetime
        now   = datetime.now()
        start = datetime.fromisoformat(start_date)
        end   = datetime.fromisoformat(end_date)

        if start < now - timedelta(minutes=5):
            return jsonify({'success': False, 'error': 'Start date must be at least 5 minutes from now.'})
        if end <= start:
            return jsonify({'success': False, 'error': 'End date must be after the start date.'})
        if (end - start).total_seconds() < 3600:
            return jsonify({'success': False, 'error': 'Election must run for at least 1 hour.'})

        connect = mysql.connector.connect(**db_config)
        cursor  = connect.cursor(dictionary=True)
        cursor.execute("SELECT id FROM elections WHERE title=%s", (title,))
        if cursor.fetchone():
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': 'An election with this title already exists.'})

        cursor.execute(
            "SELECT title FROM elections WHERE status IN ('draft','active') AND start_date < %s AND end_date > %s",
            (end_date, start_date)
        )
        overlap = cursor.fetchone()
        cursor.fetchall()  # drain remaining rows
        if overlap:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': f"Date range overlaps with existing election: '{overlap['title']}'. Elections must not overlap."})

        cursor.execute(
            "INSERT INTO elections (title, description, start_date, end_date, status, created_by) VALUES (%s, %s, %s, %s, %s, %s)",
            (title, description, start_date, end_date, 'draft', session['admin_id'])
        )
        _log(cursor, session['admin_id'], 'Create Election', None, f'Created election "{title}" running {start_date} to {end_date}')
        connect.commit()
        cursor.close()
        connect.close()
        return jsonify({'success': True})

@admin.route('/edit-election/<int:id>', methods=['POST'])
def edit_election(id):
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor  = connect.cursor(dictionary=True)
    if request.method == 'POST':
        title       = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        start_date  = request.form.get('start_date', '')
        end_date    = request.form.get('end_date', '')

        if not title:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': 'Title is required.'})
        if len(title) < 5:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': 'Title must be at least 5 characters.'})
        if len(title) > 150:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': 'Title cannot exceed 150 characters.'})
        if len(description) > 300:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': 'Description cannot exceed 300 characters.'})
        if not start_date or not end_date:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': 'Start and end dates are required.'})

        cursor.execute("SELECT status FROM elections WHERE id=%s", (id,))
        election = cursor.fetchone()
        if election and election['status'] == 'draft':
            from datetime import datetime
            now   = datetime.now()
            start = datetime.fromisoformat(start_date)
            end   = datetime.fromisoformat(end_date)
            if start < now - timedelta(minutes=5):
                cursor.close(); connect.close()
                return jsonify({'success': False, 'error': 'Start date must be at least 5 minutes from now.'})
            if end <= start:
                cursor.close(); connect.close()
                return jsonify({'success': False, 'error': 'End date must be after the start date.'})
            if (end - start).total_seconds() < 3600:
                cursor.close(); connect.close()
                return jsonify({'success': False, 'error': 'Election must run for at least 1 hour.'})

        cursor.execute(
            "SELECT title FROM elections WHERE status IN ('draft','active') AND id != %s AND start_date < %s AND end_date > %s",
            (id, end_date, start_date)
        )
        overlap = cursor.fetchone()
        cursor.fetchall()  # drain remaining rows
        if overlap:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': f"Date range overlaps with existing election: '{overlap['title']}'. Elections must not overlap."})

        cursor.execute(
            "UPDATE elections SET title=%s, description=%s, start_date=%s, end_date=%s WHERE id=%s",
            (title, description, start_date, end_date, id)
        )
        _log(cursor, session['admin_id'], 'Edit Election', id, f'Edited election {id}: title="{title}", dates={start_date} to {end_date}')
        connect.commit()
        cursor.close()
        connect.close()
        return jsonify({'success': True})
    cursor.close()
    connect.close()
    return jsonify({'success': False, 'error': 'Invalid request'})

@admin.route('/restore-election/<int:id>')
def restore_election(id):
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT is_archived FROM elections WHERE id=%s", (id,))
    election = cursor.fetchone()
    if not election or not election['is_archived']:
        cursor.close(); connect.close()
        return redirect(url_for('admin.elections') + '?error=Only archived elections can be restored.')
    cursor.execute("UPDATE elections SET is_archived=0 WHERE id=%s", (id,))
    connect.commit()
    cursor.close(); connect.close()
    return redirect(url_for('admin.elections'))

@admin.route('/start-election/<int:id>')
def start_election(id):
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT status FROM elections WHERE id=%s", (id,))
    election = cursor.fetchone()
    if not election or election['status'] != 'draft':
        cursor.close(); connect.close()
        return redirect(url_for('admin.elections') + '?error=Only draft elections can be started.')
    cursor.execute("UPDATE elections SET status='active', start_date=NOW() WHERE id=%s", (id,))
    _log(cursor, session['admin_id'], 'Start Election', id, f'Started election {id}')
    connect.commit()
    cursor.close(); connect.close()
    return redirect(url_for('admin.elections'))

@admin.route('/end-election/<int:id>')
def end_election(id):
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT status FROM elections WHERE id=%s", (id,))
    election = cursor.fetchone()
    if not election or election['status'] != 'active':
        cursor.close(); connect.close()
        return redirect(url_for('admin.elections') + '?error=Only active elections can be ended.')
    cursor.execute("UPDATE elections SET status='ended' WHERE id=%s", (id,))
    _log(cursor, session['admin_id'], 'End Election', id, f'Ended election {id}')
    connect.commit()
    cursor.close(); connect.close()
    return redirect(url_for('admin.elections'))

@admin.route('/archive-election/<int:id>')
def archive_election(id):
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT status, is_archived FROM elections WHERE id=%s", (id,))
    election = cursor.fetchone()
    if not election or election['is_archived']:
        cursor.close(); connect.close()
        return redirect(url_for('admin.elections') + '?error=Election cannot be archived.')
    cursor.execute("UPDATE elections SET is_archived=1 WHERE id=%s", (id,))
    connect.commit()
    cursor.close(); connect.close()
    return redirect(url_for('admin.elections'))

@admin.route('/delete-election/<int:id>')
def delete_election(id):
    if 'admin_id' not in session:  
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT status FROM elections WHERE id=%s", (id,))
    election = cursor.fetchone()
    if election and election['status'] == 'active':
        cursor.close()
        connect.close()
        return redirect(url_for('admin.elections') + '?error=Cannot delete an active election')
    cursor.execute("SELECT photo FROM candidates WHERE election_id=%s AND photo IS NOT NULL", (id,))
    for row in cursor.fetchall():
        photo_path = os.path.join(UPLOAD_FOLDER, row['photo'])
        if os.path.exists(photo_path):
            os.remove(photo_path)
    cursor.execute("DELETE FROM elections WHERE id=%s", (id,))
    connect.commit()
    cursor.close()
    connect.close()
    return redirect(url_for('admin.elections'))

@admin.route('/register-voter', methods=['POST'])
def register_voter():
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    from voting_system.Authentication.login import send_activation_email
    user_id = request.form.get('user_id', '').strip()
    firstname = request.form.get('firstname', '').strip()
    lastname = request.form.get('lastname', '').strip()
    email = request.form.get('email', '').strip()

    if not user_id or not firstname or not lastname or not email:
        return jsonify({'success': False, 'error': 'All fields are required.'})

    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT user_id, email FROM users WHERE user_id=%s OR email=%s", (user_id, email))
    existing = cursor.fetchall()
    if existing:
        cursor.close(); connect.close()
        dup_ids    = [r['user_id'] for r in existing]
        dup_emails = [r['email']   for r in existing]
        if user_id in dup_ids and email in dup_emails:
            return jsonify({'success': False, 'error': 'Student ID and Email are already registered.'})
        if user_id in dup_ids:
            return jsonify({'success': False, 'error': 'Student ID is already registered.'})
        return jsonify({'success': False, 'error': 'Email is already registered.'})

    code = str(random.randint(100000, 999999))
    cursor.execute(
        "INSERT INTO users (user_id, firstname, lastname, email, password, status, activation_code) VALUES (%s, %s, %s, %s, '', 'inactive', %s)",
        (user_id, firstname, lastname, email, code)
    )
    _log(cursor, session['admin_id'], 'Register Voter', None, f'Registered inactive voter {user_id} ({firstname} {lastname})')
    connect.commit()
    cursor.close(); connect.close()

    try:
        send_activation_email(email, firstname, code)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Account created but email failed: {str(e)}'})

    return jsonify({'success': True})


@admin.route('/delete-voter/<int:user_id>')
def delete_voter(user_id):
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE id=%s AND status='inactive'", (user_id,))
    voter = cursor.fetchone()
    if not voter:
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Voter not found or already active.'})
    # Guard: do not delete if voter has cast any votes (would corrupt election results)
    cursor.execute("SELECT COUNT(*) as cnt FROM votes WHERE user_id=%s", (user_id,))
    if cursor.fetchone()['cnt'] > 0:
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Cannot delete this voter — they have already cast votes in an election.'})
    cursor.execute("DELETE FROM users WHERE id=%s", (user_id,))
    connect.commit()
    cursor.close(); connect.close()
    return jsonify({'success': True})

@admin.route('/resend-activation/<int:user_id>', methods=['POST'])
def resend_activation(user_id):
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    from voting_system.Authentication.login import send_activation_email
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE id=%s AND status='inactive'", (user_id,))
    voter = cursor.fetchone()
    if not voter:
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Voter not found or already active.'})
    code = str(random.randint(100000, 999999))
    cursor.execute("UPDATE users SET activation_code=%s WHERE id=%s", (code, user_id))
    connect.commit()
    cursor.close(); connect.close()
    try:
        send_activation_email(voter['email'], voter['firstname'], code)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Email failed: {str(e)}'})
    return jsonify({'success': True})


@admin.route('/change-email/<int:user_id>', methods=['POST'])
def change_email(user_id):
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    from voting_system.Authentication.login import send_activation_email
    email = request.form.get('email', '').strip()
    if not email:
        return jsonify({'success': False, 'error': 'Email is required.'})
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE id=%s AND status='inactive'", (user_id,))
    voter = cursor.fetchone()
    if not voter:
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Voter not found or already active.'})
    cursor.execute("SELECT id FROM users WHERE email=%s AND id != %s", (email, user_id))
    if cursor.fetchone():
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Email already in use by another account.'})
    code = str(random.randint(100000, 999999))
    cursor.execute("UPDATE users SET email=%s, activation_code=%s WHERE id=%s", (email, code, user_id))
    connect.commit()
    cursor.close(); connect.close()
    try:
        send_activation_email(email, voter['firstname'], code)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Email updated but sending failed: {str(e)}'})
    return jsonify({'success': True})


@admin.route('/results-data/<int:election_id>')
def results_data(election_id):
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'})
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)

    cursor.execute("SELECT * FROM elections WHERE id=%s", (election_id,))
    election = cursor.fetchone()
    if not election:
        cursor.close(); connect.close()
        return jsonify({'error': 'Not found'})

    cursor.execute("SELECT COUNT(*) as total FROM users WHERE role='user' AND status='active'")
    total_voters = cursor.fetchone()['total']
    # Unique voters who participated (opened ballot)
    cursor.execute("SELECT COUNT(DISTINCT user_id) as voted FROM votes WHERE election_id=%s", (election_id,))
    total_participated = cursor.fetchone()['voted']
    # Total individual position votes cast
    cursor.execute("SELECT COUNT(*) as total_votes FROM votes WHERE election_id=%s", (election_id,))
    total_votes = cursor.fetchone()['total_votes']

    cursor.execute("SELECT * FROM positions WHERE election_id=%s", (election_id,))
    positions = cursor.fetchall()
    for pos in positions:
        cursor.execute("""
            SELECT c.id, c.firstname, c.lastname, c.photo,
                   COUNT(v.id) as vote_count
            FROM candidates c
            LEFT JOIN votes v ON v.candidate_id = c.id AND v.election_id=%s
            WHERE c.position_id=%s
            GROUP BY c.id ORDER BY vote_count DESC
        """, (election_id, pos['id']))
        candidates = cursor.fetchall()
        total_pos_votes = sum(c['vote_count'] for c in candidates)
        for c in candidates:
            c['percentage'] = round((c['vote_count'] / total_pos_votes * 100) if total_pos_votes > 0 else 0, 1)
        pos['candidates'] = candidates
        pos['winner'] = candidates[0] if candidates and candidates[0]['vote_count'] > 0 else None

    cursor.close(); connect.close()
    return jsonify({
        'election': {'title': election['title'], 'status': election['status']},
        'total_voters': total_voters,
        'total_participated': total_participated,
        'total_votes': total_votes,
        'turnout': round((total_participated / total_voters * 100) if total_voters > 0 else 0, 1),
        'positions': [{
            'name': pos['name'],
            'winner': f"{pos['winner']['firstname']} {pos['winner']['lastname']}" if pos['winner'] else None,
            'candidates': [{
                'id': c['id'],
                'name': f"{c['firstname']} {c['lastname']}",
                'photo': c['photo'],
                'vote_count': c['vote_count'],
                'percentage': c['percentage']
            } for c in pos['candidates']]
        } for pos in positions]
    })


@admin.route('/logs-data')
def logs_data():
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'})
    election_id = request.args.get('election_id', '')
    page = int(request.args.get('page', 1))
    per_page = 10
    offset = (page - 1) * per_page

    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("DELETE FROM activity_logs WHERE created_at < NOW() - INTERVAL 6 MONTH")
    connect.commit()

    if election_id:
        cursor.execute("""
            SELECT COUNT(*) as total FROM activity_logs l
            JOIN users u ON l.user_id = u.id
            WHERE l.election_id = %s AND u.role = 'user'
        """, (election_id,))
    else:
        cursor.execute("""
            SELECT COUNT(*) as total FROM activity_logs l
            JOIN users u ON l.user_id = u.id
            WHERE u.role = 'user'
        """)
    total = cursor.fetchone()['total']
    total_pages = max(1, -(-total // per_page))

    if election_id:
        cursor.execute("""
            SELECT l.id, u.user_id, u.firstname, u.lastname, l.action, e.title as election_title, l.details, l.created_at
            FROM activity_logs l
            JOIN users u ON l.user_id = u.id
            LEFT JOIN elections e ON l.election_id = e.id
            WHERE l.election_id = %s AND u.role = 'user'
            ORDER BY l.created_at DESC
            LIMIT %s OFFSET %s
        """, (election_id, per_page, offset))
    else:
        cursor.execute("""
            SELECT l.id, u.user_id, u.firstname, u.lastname, l.action, e.title as election_title, l.details, l.created_at
            FROM activity_logs l
            JOIN users u ON l.user_id = u.id
            LEFT JOIN elections e ON l.election_id = e.id
            WHERE u.role = 'user'
            ORDER BY l.created_at DESC
            LIMIT %s OFFSET %s
        """, (per_page, offset))
    logs = cursor.fetchall()
    cursor.close()
    connect.close()
    for log in logs:
        log['created_at'] = log['created_at'].strftime('%b %d, %Y at %I:%M %p')
    return jsonify({'logs': logs, 'total': total, 'page': page, 'total_pages': total_pages})


@admin.route('/download-voter-template')
def download_voter_template():
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Voters'
    ws.append(['student_id', 'firstname', 'lastname', 'email'])
    ws.append(['261-0001', 'Juan', 'Dela Cruz', 'juan@school.edu'])
    ws.append(['261-0002', 'Maria', 'Santos', 'maria@school.edu'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='voters_template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin.route('/import-voters', methods=['POST'])
def import_voters():
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    from voting_system.Authentication.login import send_activation_email

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'No file uploaded.'})
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload a .xlsx file.'})

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()), data_only=True)
    except Exception:
        return jsonify({'success': False, 'error': 'Could not read the file. Make sure it is a valid .xlsx file.'})

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return jsonify({'success': False, 'error': 'The file is empty.'})

    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    required = {'student_id', 'firstname', 'lastname', 'email'}
    if not required.issubset(set(headers)):
        return jsonify({'success': False, 'error': 'Invalid file format. Expected columns: student_id, firstname, lastname, email.'})

    si = headers.index('student_id')
    fi = headers.index('firstname')
    li = headers.index('lastname')
    ei = headers.index('email')

    data_rows = rows[1:]
    if not data_rows or all(r[si] is None and r[fi] is None and r[li] is None and r[ei] is None for r in data_rows):
        return jsonify({'success': False, 'error': 'No voters found in the file.'})

    if len(data_rows) > 30:
        return jsonify({'success': False, 'error': 'Too many rows. Maximum is 30 voters per import.'})

    import re
    NAME_REGEX  = re.compile(r"^[A-Za-z\s'\-]{3,30}$")
    EMAIL_REGEX = re.compile(r"^(?!.*\.\.)(?!.*\.$)[^\W][a-zA-Z0-9._%+\-]{0,63}@[a-zA-Z0-9\-]+(\.[a-zA-Z0-9\-]+)*\.[a-zA-Z]{2,}$")
    ID_REGEX    = re.compile(r"^\d{3}-\d{4}$")

    connect = mysql.connector.connect(**db_config)
    cursor  = connect.cursor(dictionary=True)

    added   = 0
    skipped = []

    for i, row in enumerate(data_rows, start=2):
        student_id = str(row[si]).strip() if row[si] is not None else ''
        firstname  = str(row[fi]).strip().title() if row[fi] is not None else ''
        lastname   = str(row[li]).strip().title() if row[li] is not None else ''
        email      = str(row[ei]).strip().lower() if row[ei] is not None else ''

        if not student_id or not firstname or not lastname or not email:
            skipped.append(f'Row {i}: missing required fields.')
            continue
        if not ID_REGEX.match(student_id):
            skipped.append(f'Row {i}: invalid student_id "{student_id}". Format must be XXX-XXXX.')
            continue
        if not NAME_REGEX.match(firstname):
            skipped.append(f'Row {i}: invalid firstname "{firstname}".')
            continue
        if not NAME_REGEX.match(lastname):
            skipped.append(f'Row {i}: invalid lastname "{lastname}".')
            continue
        if not EMAIL_REGEX.match(email):
            skipped.append(f'Row {i}: invalid email "{email}".')
            continue

        cursor.execute("SELECT user_id, email FROM users WHERE user_id=%s OR email=%s", (student_id, email))
        existing = cursor.fetchall()
        if existing:
            dup_ids    = [r['user_id'] for r in existing]
            dup_emails = [r['email']   for r in existing]
            if student_id in dup_ids and email in dup_emails:
                skipped.append(f'Row {i}: student_id and email already registered.')
            elif student_id in dup_ids:
                skipped.append(f'Row {i}: student_id "{student_id}" already registered.')
            else:
                skipped.append(f'Row {i}: email "{email}" already registered.')
            continue

        code = str(random.randint(100000, 999999))
        cursor.execute(
            "INSERT INTO users (user_id, firstname, lastname, email, password, status, activation_code) VALUES (%s, %s, %s, %s, '', 'inactive', %s)",
            (student_id, firstname, lastname, email, code)
        )
        connect.commit()

        # Send activation email in background thread to avoid blocking the response
        import threading
        def send_email_safe():
            try:
                send_activation_email(email, firstname, code)
            except Exception as e:
                print(f'[EMAIL ERROR] Failed to send activation email to {email}: {e}')
        
        threading.Thread(target=send_email_safe, daemon=True).start()
        added += 1

    cursor.close()
    connect.close()

    msg = f'{added} voter(s) registered'
    if skipped:
        msg += f', {len(skipped)} skipped: ' + ' | '.join(skipped)
    return jsonify({'success': True, 'message': msg})



@admin.route('/download-candidate-template')
def download_candidate_template():
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Candidates'
    ws.append(['firstname', 'lastname', 'position'])
    ws.append(['Juan', 'Dela Cruz', 'President'])
    ws.append(['Maria', 'Santos', 'Vice President'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='candidates_template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin.route('/import-candidates', methods=['POST'])
def import_candidates():
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    election_id = request.form.get('election_id')
    file = request.files.get('file')

    if not election_id:
        return jsonify({'success': False, 'error': 'Please select an election.'})
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'No file uploaded.'})
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload a .xlsx file.'})

    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)

    cursor.execute("SELECT status FROM elections WHERE id=%s", (election_id,))
    election = cursor.fetchone()
    if not election or election['status'] == 'ended':
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Cannot import candidates to an ended election.'})

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()), data_only=True)
    except Exception:
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Could not read the file. Make sure it is a valid .xlsx file.'})

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'The file is empty.'})

    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    required = {'firstname', 'lastname', 'position'}
    if not required.issubset(set(headers)):
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Invalid file format. Expected columns: firstname, lastname, position.'})

    fi = headers.index('firstname')
    li = headers.index('lastname')
    pi = headers.index('position')

    data_rows = rows[1:]
    if not data_rows or all(r[fi] is None and r[li] is None and r[pi] is None for r in data_rows):
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'No candidates found in the file.'})

    if len(data_rows) > 30:
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Too many rows. Maximum is 30 candidates per import.'})

    cursor.execute("SELECT id, LOWER(name) as name FROM positions WHERE election_id=%s", (election_id,))
    positions_map = {row['name'].replace('\xa0', ' ').strip(): row['id'] for row in cursor.fetchall()}

    import re
    NAME_REGEX = re.compile(r"^[A-Za-z\s'\-]{3,30}$")

    added = 0
    skipped = []

    for i, row in enumerate(data_rows, start=2):
        firstname = str(row[fi]).strip().title() if row[fi] is not None else ''
        lastname  = str(row[li]).strip().title() if row[li] is not None else ''
        position  = str(row[pi]).strip().replace('\xa0', ' ').strip() if row[pi] is not None else ''

        if not firstname or not lastname or not position:
            skipped.append(f'Row {i}: missing required fields.')
            continue
        if not NAME_REGEX.match(firstname):
            skipped.append(f'Row {i}: invalid firstname "{firstname}".')
            continue
        if not NAME_REGEX.match(lastname):
            skipped.append(f'Row {i}: invalid lastname "{lastname}".')
            continue

        pos_id = positions_map.get(position.lower())
        if not pos_id:
            skipped.append(f'Row {i}: position "{position}" not found. Available: {", ".join(positions_map.keys()) or "none — add positions first"}')
            continue

        cursor.execute(
            "SELECT id FROM candidates WHERE firstname=%s AND lastname=%s AND election_id=%s",
            (firstname, lastname, election_id)
        )
        if cursor.fetchone():
            skipped.append(f'Row {i}: "{firstname} {lastname}" is a duplicate.')
            continue

        cursor.execute(
            "INSERT INTO candidates (firstname, lastname, photo, position_id, election_id) VALUES (%s, %s, NULL, %s, %s)",
            (firstname, lastname, pos_id, election_id)
        )
        added += 1

    connect.commit()
    cursor.close()
    connect.close()

    msg = f'{added} candidate(s) added'
    if skipped:
        msg += f', {len(skipped)} skipped: ' + ' | '.join(skipped)
    return jsonify({'success': True, 'message': msg})


# ── POSITIONS ──

@admin.route('/add-position', methods=['POST'])
def add_position():
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    name = request.form.get('name', '').strip()
    election_id = request.form.get('election_id')
    if not name or not election_id:
        return jsonify({'success': False, 'error': 'All fields are required.'})
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT status FROM elections WHERE id=%s", (election_id,))
    election = cursor.fetchone()
    if not election or election['status'] == 'ended':
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Cannot add a position to an ended election.'})
    cursor.execute("INSERT INTO positions (name, election_id) VALUES (%s, %s)", (name, election_id))
    connect.commit()
    new_id = cursor.lastrowid
    cursor.close()
    connect.close()
    return jsonify({'success': True, 'id': new_id, 'name': name})


@admin.route('/delete-position/<int:id>')
def delete_position(id):
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT photo FROM candidates WHERE position_id=%s AND photo IS NOT NULL", (id,))
    for row in cursor.fetchall():
        photo_path = os.path.join(UPLOAD_FOLDER, row['photo'])
        if os.path.exists(photo_path):
            os.remove(photo_path)
    cursor.execute("DELETE FROM positions WHERE id=%s", (id,))
    connect.commit()
    cursor.close()
    connect.close()
    return jsonify({'success': True})


@admin.route('/get-positions/<int:election_id>')
def get_positions(election_id):
    if 'admin_id' not in session:
        return jsonify([])
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT * FROM positions WHERE election_id=%s", (election_id,))
    positions = cursor.fetchall()
    cursor.close()
    connect.close()
    return jsonify(positions)


# ── CANDIDATES ──

@admin.route('/get-candidates/<int:election_id>')
def get_candidates(election_id):
    if 'admin_id' not in session:
        return jsonify([])
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("""
        SELECT c.*, p.name as position_name
        FROM candidates c
        JOIN positions p ON c.position_id = p.id
        WHERE c.election_id=%s
        ORDER BY p.name, c.lastname
    """, (election_id,))
    candidates = cursor.fetchall()
    cursor.close()
    connect.close()
    return jsonify(candidates)


@admin.route('/add-candidate', methods=['POST'])
def add_candidate():
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    firstname = request.form.get('firstname', '').strip()
    lastname = request.form.get('lastname', '').strip()
    position_id = request.form.get('position_id')
    election_id = request.form.get('election_id')
    if not firstname or not lastname or not position_id or not election_id:
        return jsonify({'success': False, 'error': 'All fields are required.'})

    connect = mysql.connector.connect(**db_config)
    cursor  = connect.cursor(dictionary=True)
    cursor.execute("SELECT status FROM elections WHERE id=%s", (election_id,))
    election_check = cursor.fetchone()
    if not election_check or election_check['status'] == 'ended':
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': 'Cannot add a candidate to an ended election.'})

    cursor.execute(
        "SELECT id FROM candidates WHERE firstname=%s AND lastname=%s AND election_id=%s",
        (firstname, lastname, election_id)
    )
    if cursor.fetchone():
        cursor.close(); connect.close()
        return jsonify({'success': False, 'error': f'{firstname} {lastname} is already registered as a candidate in this election.'})

    photo_filename = None
    file = request.files.get('photo')
    if file and file.filename and allowed_file(file.filename):
        import uuid
        ext = file.filename.rsplit('.', 1)[1].lower()
        photo_filename = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(UPLOAD_FOLDER, photo_filename))

    cursor.execute(
        "INSERT INTO candidates (firstname, lastname, photo, position_id, election_id) VALUES (%s, %s, %s, %s, %s)",
        (firstname, lastname, photo_filename, position_id, election_id)
    )
    connect.commit()
    new_id = cursor.lastrowid
    cursor.close()
    connect.close()
    return jsonify({'success': True, 'id': new_id})


@admin.route('/edit-candidate/<int:id>', methods=['POST'])
def edit_candidate(id):
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    firstname = request.form.get('firstname', '').strip()
    lastname = request.form.get('lastname', '').strip()
    position_id = request.form.get('position_id')
    if not firstname or not lastname or not position_id:
        return jsonify({'success': False, 'error': 'All fields are required.'})

    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)

    cursor.execute("SELECT election_id FROM candidates WHERE id=%s", (id,))
    row = cursor.fetchone()
    if row:
        cursor.execute(
            "SELECT id FROM candidates WHERE firstname=%s AND lastname=%s AND election_id=%s AND id != %s",
            (firstname, lastname, row['election_id'], id)
        )
        if cursor.fetchone():
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': f'{firstname} {lastname} is already registered as a candidate in this election.'})

    file = request.files.get('photo')
    if file and file.filename and allowed_file(file.filename):
        cursor.execute("SELECT photo FROM candidates WHERE id=%s", (id,))
        old = cursor.fetchone()
        if old and old['photo']:
            old_path = os.path.join(UPLOAD_FOLDER, old['photo'])
            if os.path.exists(old_path):
                os.remove(old_path)
        import uuid
        ext = file.filename.rsplit('.', 1)[1].lower()
        new_filename = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(UPLOAD_FOLDER, new_filename))
        cursor.execute(
            "UPDATE candidates SET firstname=%s, lastname=%s, position_id=%s, photo=%s WHERE id=%s",
            (firstname, lastname, position_id, new_filename, id)
        )
    else:
        cursor.execute(
            "UPDATE candidates SET firstname=%s, lastname=%s, position_id=%s WHERE id=%s",
            (firstname, lastname, position_id, id)
        )
    connect.commit()
    cursor.close()
    connect.close()
    return jsonify({'success': True})


@admin.route('/users')
def list_users():
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("""
        SELECT id, user_id, firstname, lastname, email, status, role
        FROM users
        WHERE role = 'user'
        ORDER BY firstname, lastname
    """)
    users = cursor.fetchall()
    cursor.close()
    connect.close()
    return render_template('admin_users.html', username=session.get('admin_name'), users=users, active_page='list_users')


@admin.route('/delete-candidate/<int:id>')
def delete_candidate(id):
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    connect = mysql.connector.connect(**db_config)
    cursor = connect.cursor(dictionary=True)
    cursor.execute("SELECT photo FROM candidates WHERE id=%s", (id,))
    candidate = cursor.fetchone()
    if candidate and candidate['photo']:
        photo_path = os.path.join(UPLOAD_FOLDER, candidate['photo'])
        if os.path.exists(photo_path):
            os.remove(photo_path)
    cursor.execute("DELETE FROM candidates WHERE id=%s", (id,))
    connect.commit()
    cursor.close()
    connect.close()
    return jsonify({'success': True})


# ── ADMIN PROFILE ──

ADMIN_PHOTO_FOLDER = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'static', 'profile_photos')

@admin.route('/profile', methods=['GET', 'POST'])
def profile():
    if 'admin_id' not in session:
        return redirect(url_for('auth.login'))

    connect = mysql.connector.connect(**db_config)
    cursor  = connect.cursor(dictionary=True)

    if request.method == 'POST':
        import bcrypt
        current_pw = request.form.get('current_password', '')
        new_pw     = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')

        cursor.execute("SELECT * FROM users WHERE id=%s", (session['admin_id'],))
        admin_data = cursor.fetchone()
        if not admin_data:
            cursor.close(); connect.close()
            session.clear()
            return jsonify({'success': False, 'error': 'Session expired. Please log in again.'})
        stored = admin_data['password']
        if stored.startswith('$2b$') or stored.startswith('$2a$'):
            match = bcrypt.checkpw(current_pw.encode('utf-8'), stored.encode('utf-8'))
        else:
            match = (current_pw == stored)

        if not match:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': 'Current password is incorrect.'})
        if new_pw != confirm_pw:
            cursor.close(); connect.close()
            return jsonify({'success': False, 'error': 'New passwords do not match.'})
        if stored.startswith('$2b$') or stored.startswith('$2a$'):
            if bcrypt.checkpw(new_pw.encode('utf-8'), stored.encode('utf-8')):
                cursor.close(); connect.close()
                return jsonify({'success': False, 'error': 'New password cannot be the same as your current password.'})

        hashed = bcrypt.hashpw(new_pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        cursor.execute("UPDATE users SET password=%s WHERE id=%s", (hashed, session['admin_id']))
        connect.commit()
        cursor.close(); connect.close()
        return jsonify({'success': True})

    cursor.execute("SELECT id, user_id, firstname, lastname, email, profile_photo FROM users WHERE id=%s", (session['admin_id'],))
    admin_data = cursor.fetchone()
    if not admin_data:
        cursor.close(); connect.close()
        session.clear()
        return redirect(url_for('auth.login') + '?error=Your session is no longer valid. Please log in again.')
    cursor.execute("SELECT is_enabled FROM user_2fa WHERE user_id=%s", (session['admin_id'],))
    twofa = cursor.fetchone()
    admin_data['twofa_enabled'] = twofa['is_enabled'] if twofa else False

    # Stats
    cursor.execute("SELECT COUNT(*) as total FROM elections WHERE created_by=%s", (session['admin_id'],))
    admin_data['elections_managed'] = cursor.fetchone()['total']
    cursor.execute("SELECT COUNT(*) as total FROM users WHERE role='user' AND status='active'")
    admin_data['voters_registered'] = cursor.fetchone()['total']

    # Recent activity
    cursor.execute("""
        SELECT action, details, created_at FROM activity_logs
        WHERE user_id=%s ORDER BY created_at DESC LIMIT 5
    """, (session['admin_id'],))
    recent_logs = cursor.fetchall()

    cursor.close(); connect.close()
    return render_template('admin_profile.html', username=session.get('admin_name'), admin=admin_data, recent_logs=recent_logs, active_page='profile')


@admin.route('/upload-admin-photo', methods=['POST'])
def upload_admin_photo():
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    file = request.files.get('photo')
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'No file selected.'})
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Only image files are allowed (jpg, png, gif).'})

    file.seek(0, 2)
    size_mb = file.tell() / (1024 * 1024)
    file.seek(0)
    if size_mb > 5:
        return jsonify({'success': False, 'error': f'File must be under 5MB (current: {size_mb:.1f}MB).'})

    import uuid
    connect = mysql.connector.connect(**db_config)
    cursor  = connect.cursor(dictionary=True)
    cursor.execute("SELECT profile_photo FROM users WHERE id=%s", (session['admin_id'],))
    row = cursor.fetchone()
    if row and row['profile_photo']:
        old_path = os.path.join(ADMIN_PHOTO_FOLDER, row['profile_photo'])
        if os.path.exists(old_path):
            os.remove(old_path)

    ext      = file.filename.rsplit('.', 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    file.save(os.path.join(ADMIN_PHOTO_FOLDER, filename))

    cursor.execute("UPDATE users SET profile_photo=%s WHERE id=%s", (filename, session['admin_id']))
    connect.commit()
    cursor.close(); connect.close()
    return jsonify({'success': True, 'filename': filename})


@admin.route('/delete-admin-photo', methods=['POST'])
def delete_admin_photo():
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    connect = mysql.connector.connect(**db_config)
    cursor  = connect.cursor(dictionary=True)
    cursor.execute("SELECT profile_photo FROM users WHERE id=%s", (session['admin_id'],))
    row = cursor.fetchone()
    if row and row['profile_photo']:
        old_path = os.path.join(ADMIN_PHOTO_FOLDER, row['profile_photo'])
        if os.path.exists(old_path):
            os.remove(old_path)
        cursor.execute("UPDATE users SET profile_photo=NULL WHERE id=%s", (session['admin_id'],))
        connect.commit()
    cursor.close(); connect.close()
    return jsonify({'success': True})


@admin.route('/admin-profile-photo/<filename>')
def admin_profile_photo(filename):
    return send_from_directory(ADMIN_PHOTO_FOLDER, filename)


@admin.route('/toggle-admin-2fa/<action>', methods=['POST'])
def toggle_admin_2fa(action):
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    if action not in ('enable', 'disable'):
        return jsonify({'success': False, 'error': 'Invalid action.'})
    connect = mysql.connector.connect(**db_config)
    cursor  = connect.cursor(dictionary=True)
    cursor.execute("SELECT id FROM user_2fa WHERE user_id=%s", (session['admin_id'],))
    existing = cursor.fetchone()
    enabled  = 1 if action == 'enable' else 0
    if existing:
        cursor.execute("UPDATE user_2fa SET is_enabled=%s, otp_code=NULL, otp_expires=NULL WHERE user_id=%s", (enabled, session['admin_id']))
    else:
        cursor.execute("INSERT INTO user_2fa (user_id, is_enabled) VALUES (%s, %s)", (session['admin_id'], enabled))
    connect.commit()
    cursor.close(); connect.close()
    return jsonify({'success': True})
